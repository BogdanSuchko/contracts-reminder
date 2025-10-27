from __future__ import annotations

import asyncio
import re
from datetime import datetime, timedelta
from logging import Logger
from pathlib import Path
from typing import Optional

import requests
from openpyxl import load_workbook
from zoneinfo import ZoneInfo

from contract_bot.config import AppConfig
from contract_bot.service.reminder import ReminderService
from contract_bot.storage.file_repository import FileRepository
from contract_bot.storage.state_store import StateStore


class SheetSyncService:
    def __init__(
        self,
        config: AppConfig,
        file_repository: FileRepository,
        state_store: StateStore,
        logger: Logger,
    ) -> None:
        self._config = config
        self._file_repository = file_repository
        self._state_store = state_store
        self._logger = logger
        self._timezone = ZoneInfo(config.scheduler.timezone)
        self._last_sync: Optional[datetime] = None
        self._interval = timedelta(minutes=config.integrations.sheet_sync_interval_minutes)
        self._tolerance = timedelta(seconds=5)
        self._reminder_service: Optional[ReminderService] = None
        self._current_reminder_days = config.scheduler.reminder_days

    def set_reminder_service(self, service: ReminderService) -> None:
        self._reminder_service = service
        service.update_reminder_days(self._current_reminder_days)

    @property
    def current_reminder_days(self) -> int:
        return self._current_reminder_days

    @property
    def enabled(self) -> bool:
        return bool(self._config.integrations.google_sheet_id)

    async def sync(self, *, force: bool = False) -> bool:
        if not self.enabled:
            return False

        now = datetime.now(self._timezone)
        if (
            not force
            and self._last_sync
            and now - self._last_sync < max(self._interval - self._tolerance, timedelta())
        ):
            return False

        try:
            await asyncio.to_thread(self._download)
            self._last_sync = datetime.now(self._timezone)
            return True
        except Exception as exc:  # noqa: BLE001
            self._logger.warning("Не удалось синхронизировать Google Sheet: %s", exc)
            return False

    def _download(self) -> None:
        sheet_id = self._config.integrations.google_sheet_id
        gid = self._config.integrations.google_sheet_gid or "0"
        filename = self._config.integrations.google_sheet_filename

        urls = [
            f"https://docs.google.com/spreadsheets/d/{sheet_id}/export?format=xlsx&id={sheet_id}",
            f"https://docs.google.com/spreadsheets/d/{sheet_id}/export?format=xlsx&gid={gid}",
            f"https://docs.google.com/spreadsheets/d/{sheet_id}/export?format=csv&gid={gid}",
        ]

        last_error: Exception | None = None
        for url in urls:
            try:
                response = requests.get(url, timeout=30)
                response.raise_for_status()

                content = response.content
                path: Path
                if url.endswith("format=csv&gid={gid}".format(gid=gid)):
                    import io
                    import pandas as pd

                    df = pd.read_csv(io.BytesIO(content))
                    buffer = io.BytesIO()
                    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
                        df.to_excel(writer, index=False)
                    content = buffer.getvalue()

                path = self._file_repository.save_latest(content, filename)
                self._state_store.set_last_upload_for_all(filename)
                self._update_reminder_days(path)
                self._logger.info("Google Sheet синхронизирован по адресу %s", url)
                return
            except Exception as exc:  # noqa: BLE001
                last_error = exc
                continue

        if last_error:
            raise last_error

    def _update_reminder_days(self, path: Path) -> None:
        sheet_name_target = self._config.integrations.google_sheet_name
        try:
            wb = load_workbook(path, read_only=True, data_only=True)
            try:
                sheet = wb[sheet_name_target] if sheet_name_target in wb.sheetnames else wb.active
                raw_value = sheet["G5"].value or sheet["F5"].value
            finally:
                wb.close()
        except Exception as exc:  # noqa: BLE001
            self._logger.warning("Не удалось прочитать горизонт напоминаний: %s", exc)
            return

        days = self._parse_reminder_days(raw_value)
        if days is None:
            return

        self._current_reminder_days = days
        if self._reminder_service is not None:
            self._reminder_service.update_reminder_days(days)

    def _parse_reminder_days(self, raw: object) -> int | None:
        if raw is None:
            return None
        if isinstance(raw, (int, float)):
            value = int(raw)
            return value if value > 0 else None

        text = str(raw).strip().lower()
        if not text:
            return None

        months = 0
        days = 0
        match_month = re.search(r"(\d+)\s*мес", text)
        if match_month:
            months = int(match_month.group(1))
        match_day = re.search(r"(\d+)\s*д", text)
        if match_day:
            days = int(match_day.group(1))

        total = months * 30 + days
        if total <= 0:
            return None
        return total
